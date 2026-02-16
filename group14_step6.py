import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict

# =============================================================================
# Load workbook
# =============================================================================
FILENAME = "2BSCS-A _ Tossed Coin Raw Data.xlsx"
wb = openpyxl.load_workbook(FILENAME, data_only=True)

# =============================================================================
# SURFACE ASSIGNMENTS
# =============================================================================
SURFACE_MAP = {
    "GROUP 1": "Wood", "GROUP 2": "Wood", "GROUP 3": "Wood", "GROUP 4": "Wood",
    "GROUP 5": "Wood", "GROUP 6": "Wood", "GROUP 7": "Wood", "GROUP 8": "Wood",
    "GROUP 9": "Tiles", "GROUP 10": "Tiles", "GROUP 11": "Tiles",
    "GROUP 12": "Tiles", "GROUP 13": "Tiles",
    "GROUP 14": "Tiles",
    "GROUP 15": "Tiles",
}

# =============================================================================
# PARSE — extract individual toss results (H=1, T=0)
# =============================================================================
all_records = []


def derive_from_cumulative(cum_h_list):
    """Given a list of cumulative-H values, return per-toss H (1/0)."""
    results = []
    prev = 0
    for ch in cum_h_list:
        if ch is None:
            ch = prev
        results.append(1 if int(ch) > prev else 0)
        prev = int(ch)
    return results


def safe_int(v, default=0):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def add(group, coin, tosses):
    all_records.append((group, coin, SURFACE_MAP[group], tosses))


# ── GROUP 1  ── 1B & 2 Peso
ws = wb["GROUP 1"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 1", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 1", "Arabian 2", derive_from_cumulative([rows[i][6] for i in range(2, 102)]))

# ── GROUP 2  ── 1B & 5A
ws = wb["GROUP 2"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 2", "1B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 2", "5A", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 3  ── 1B (new) & 10A (old)
ws = wb["GROUP 3"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 3", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 3", "10A", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# ── GROUP 4  ── 5A & 5B
ws = wb["GROUP 4"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 4", "5A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 4", "5B", derive_from_cumulative([rows[i][8] for i in range(2, 102)]))

# ── GROUP 5  ── 1A & 1B
ws = wb["GROUP 5"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 5", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 5", "1B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 6  ── 5B & 20 Peso
ws = wb["GROUP 6"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 6", "5B", derive_from_cumulative([rows[i][3] for i in range(3, 103)]))
add("GROUP 6", "20 Peso", derive_from_cumulative([rows[i][8] for i in range(3, 103)]))

# ── GROUP 7  ── 5A & 10A
ws = wb["GROUP 7"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 7", "5A", derive_from_cumulative([rows[i][4] for i in range(1, 101)]))
add("GROUP 7", "10A", derive_from_cumulative([rows[i][12] for i in range(1, 101)]))

# ── GROUP 8  ── 1A & 10B
ws = wb["GROUP 8"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 8", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 8", "10B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 9  ── 5B (New) & 1B (New) & 20
ws = wb["GROUP 9"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 9", "5B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 9", "1B", derive_from_cumulative([rows[i][7] for i in range(2, 102)]))
add("GROUP 9", "20 Peso", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 10 ── 5B & 10B
ws = wb["GROUP 10"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 10", "5B", derive_from_cumulative([rows[i][3] for i in range(7, 107)]))
add("GROUP 10", "10B", derive_from_cumulative([rows[i][10] for i in range(7, 107)]))

# ── GROUP 11 ── 1A & 10B
ws = wb["GROUP 11"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 11", "1A", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 11", "10B", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 12 ── 5B (New) & 5A (Old)
ws = wb["GROUP 12"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 12", "5B", derive_from_cumulative([rows[i][4] for i in range(3, 103)]))
add("GROUP 12", "5A", derive_from_cumulative([rows[i][12] for i in range(3, 103)]))

# ── GROUP 13 ── 1A & 10A
ws = wb["GROUP 13"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 13", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 13", "10A", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 14 ── 1A & 20 Peso
ws = wb["GROUP 14"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 14", "1A", [safe_int(rows[i][1]) for i in range(2, 102)])
add("GROUP 14", "20 Peso", [safe_int(rows[i][3]) for i in range(2, 102)])

# ── GROUP 15 ── 1B & 5B
ws = wb["GROUP 15"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 15", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 15", "5B", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# =============================================================================
# Normalise coin class names & aggregate
# =============================================================================
def normalise(name):
    n = name.strip().upper()
    for rem in ["(NEW)", "(OLD)", "PESO", "₱"]:
        n = n.replace(rem, "")
    return n.strip()


class_tosses = defaultdict(list)
class_groups = defaultdict(list)

for grp, coin, surf, tosses in all_records:
    key = normalise(coin)
    class_tosses[key].extend(tosses)
    class_groups[key].append(f"G{grp.split()[-1]}({surf[0]})")

sorted_classes = sorted(class_tosses.keys())

# Grand combined
all_tosses_combined = []
for tosses in class_tosses.values():
    all_tosses_combined.extend(tosses)

# By surface
surface_tosses = defaultdict(list)
for _, _, surf, tosses in all_records:
    surface_tosses[surf].extend(tosses)

grand_h = sum(all_tosses_combined)
grand_t = len(all_tosses_combined) - grand_h

# =============================================================================
# Helper
# =============================================================================
def cumulative_ht(tosses):
    cum_h, cum_t = [], []
    h, t = 0, 0
    for v in tosses:
        if v == 1:
            h += 1
        else:
            t += 1
        cum_h.append(h)
        cum_t.append(t)
    return cum_h, cum_t


# ── Colour palette
BLUE = "#2563EB"
RED = "#DC2626"
DARK_BG = "#1E293B"
CARD_BG = "#334155"
TEXT_CLR = "white"

# =============================================================================
# STEP 6 — Canvas H & T (Combined) — Cumulative + Table
# =============================================================================
fig6, ax6 = plt.subplots(figsize=(12, 10))
fig6.patch.set_facecolor(DARK_BG)
ax6.set_facecolor(CARD_BG)
fig6.suptitle("Step 6 — Canvas H & T (Combined)",
              fontsize=18, fontweight="bold", color=TEXT_CLR)

total_all = len(all_tosses_combined)
cum_h_comb, cum_t_comb = cumulative_ht(all_tosses_combined)
x_comb = np.arange(1, total_all + 1)

ax6.plot(x_comb, cum_h_comb, color=BLUE, linewidth=2, label="Heads (H)")
ax6.plot(x_comb, cum_t_comb, color=RED, linewidth=2, label="Tails (T)")
ax6.set_xlabel("Toss # (all coins concatenated)", fontsize=12, color=TEXT_CLR, labelpad=10)
ax6.set_ylabel("Cumulative Count", fontsize=12, color=TEXT_CLR)
ax6.tick_params(colors=TEXT_CLR)
ax6.legend(fontsize=12)
ax6.grid(True, alpha=0.15, color="white")

ax6.annotate(f"H = {grand_h}", xy=(total_all, grand_h),
             fontsize=12, color=BLUE, fontweight="bold",
             xytext=(5, 5), textcoords="offset points")
ax6.annotate(f"T = {grand_t}", xy=(total_all, grand_t),
             fontsize=12, color=RED, fontweight="bold",
             xytext=(5, -15), textcoords="offset points")

# Summary table with surface breakdown -- built dynamically
table_data6 = [
    ["Heads (H)", str(grand_h), f"{grand_h / total_all * 100:.1f}%"],
    ["Tails (T)", str(grand_t), f"{grand_t / total_all * 100:.1f}%"],
    ["Total", str(total_all), "100.0%"],
    ["", "", ""],
]
for surf in sorted(surface_tosses.keys()):
    st = surface_tosses[surf]
    if len(st) == 0:
        continue
    sh = sum(st)
    sn = len(st)
    table_data6.append([surf, f"H={sh} T={sn-sh}", f"{sh/sn*100:.1f}% H"])
tbl6 = ax6.table(cellText=table_data6,
                 colLabels=["Outcome", "Count", "Probability"],
                 cellLoc="center", loc="bottom",
                 bbox=[0.20, -0.55, 0.60, 0.35])
tbl6.auto_set_font_size(False)
tbl6.set_fontsize(10)
for (row, col), cell in tbl6.get_celld().items():
    cell.set_edgecolor("white")
    if row == 0:
        cell.set_facecolor("#0F172A")
        cell.set_text_props(color="white", fontweight="bold")
    elif row == 4:  # separator
        cell.set_facecolor(DARK_BG)
        cell.set_edgecolor(DARK_BG)
        cell.set_text_props(color=DARK_BG)
    elif row >= 5:  # surface rows
        cell.set_facecolor("#1E3A5F")
        cell.set_text_props(color="#CBD5E1")
    else:
        cell.set_facecolor("#1E293B")
        cell.set_text_props(color="white")

plt.subplots_adjust(bottom=0.32)
plt.tight_layout(rect=[0, 0.20, 1, 0.93])
plt.savefig("group14_step6.png", dpi=150, bbox_inches="tight",
            facecolor=fig6.get_facecolor())
# plt.show()

print("\n✅  Step 6 chart saved: step6_canvas_ht_combined.png")

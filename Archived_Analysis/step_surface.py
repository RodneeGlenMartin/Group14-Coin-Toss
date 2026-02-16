import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict

FILENAME = "2BSCS-A _ Tossed Coin Raw Data.xlsx"
wb = openpyxl.load_workbook(FILENAME, data_only=True)

# Corrected Surface Map
SURFACE_MAP = {
    "GROUP 1": "Wood", "GROUP 2": "Wood", "GROUP 3": "Wood", "GROUP 4": "Wood",
    "GROUP 5": "Wood", "GROUP 6": "Wood", "GROUP 7": "Wood", "GROUP 8": "Wood",
    "GROUP 9": "Tiles", "GROUP 10": "Tiles", "GROUP 11": "Tiles",
    "GROUP 12": "Tiles", "GROUP 13": "Tiles",
    "GROUP 14": "Tiles",
    "GROUP 15": "Tiles",
}

# Colors
BLUE = "#2563EB"
RED = "#DC2626"
DARK_BG = "#1E293B"
CARD_BG = "#334155"
TEXT_CLR = "white"
SURFACE_COLORS = {"Wood": "#3B82F6", "Tiles": "#F97316"}

all_records = []

def derive_from_cumulative(cum_h_list):
    results = []
    prev = 0
    for ch in cum_h_list:
        if ch is None:
            ch = prev
        results.append(1 if int(ch) > prev else 0)
        prev = int(ch)
    return results

def safe_int(v, default=0):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default

def add(group, coin, tosses):
    all_records.append((group, coin, SURFACE_MAP[group], tosses))

# ── GROUP 1
ws = wb["GROUP 1"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 1", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 1", "Arabian 2", derive_from_cumulative([rows[i][6] for i in range(2, 102)]))

# ── GROUP 2
ws = wb["GROUP 2"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 2", "1B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 2", "5A", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 3
ws = wb["GROUP 3"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 3", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 3", "10A", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# ── GROUP 4
ws = wb["GROUP 4"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 4", "5A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 4", "5B", derive_from_cumulative([rows[i][8] for i in range(2, 102)]))

# ── GROUP 5
ws = wb["GROUP 5"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 5", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 5", "1B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 6
ws = wb["GROUP 6"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 6", "5B", derive_from_cumulative([rows[i][3] for i in range(3, 103)]))
add("GROUP 6", "20 Peso", derive_from_cumulative([rows[i][8] for i in range(3, 103)]))

# ── GROUP 7
ws = wb["GROUP 7"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 7", "5A", derive_from_cumulative([rows[i][4] for i in range(1, 101)]))
add("GROUP 7", "10A", derive_from_cumulative([rows[i][12] for i in range(1, 101)]))

# ── GROUP 8
ws = wb["GROUP 8"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 8", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 8", "10B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 9
ws = wb["GROUP 9"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 9", "5B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 9", "1B", derive_from_cumulative([rows[i][7] for i in range(2, 102)]))
add("GROUP 9", "20 Peso", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 10
ws = wb["GROUP 10"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 10", "5B", derive_from_cumulative([rows[i][3] for i in range(7, 107)]))
add("GROUP 10", "10B", derive_from_cumulative([rows[i][10] for i in range(7, 107)]))

# ── GROUP 11
ws = wb["GROUP 11"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 11", "1A", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 11", "10B", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 12
ws = wb["GROUP 12"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 12", "5B", derive_from_cumulative([rows[i][4] for i in range(3, 103)]))
add("GROUP 12", "5A", derive_from_cumulative([rows[i][12] for i in range(3, 103)]))

# ── GROUP 13
ws = wb["GROUP 13"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 13", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 13", "10A", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 14
ws = wb["GROUP 14"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 14", "1A", [safe_int(rows[i][1]) for i in range(2, 102)])
add("GROUP 14", "20 Peso", [safe_int(rows[i][3]) for i in range(2, 102)])

# ── GROUP 15
ws = wb["GROUP 15"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 15", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 15", "5B", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))


# ── Aggregate by Surface
surface_tosses = defaultdict(list)
for _, _, surf, tosses in all_records:
    surface_tosses[surf].extend(tosses)

def cumulative_ht(tosses):
    cum_h, cum_t = [], []
    h, t = 0, 0
    for v in tosses:
        if v == 1:
            h += 1
        else:
            t += 1
        cum_h.append(h)
        cum_t.append(t)
    return cum_h, cum_t


# =============================================================================
# SURFACE COMPARISON CHARTS
# =============================================================================

# ── Chart A: Cumulative H & T per surface ─────────────────────────────────
surface_order = sorted([s for s in surface_tosses if len(surface_tosses[s]) > 0])
n_surfaces = len(surface_order)
fig_s1, axes_s1 = plt.subplots(1, n_surfaces, figsize=(9 * n_surfaces, 6))
if n_surfaces == 1:
    axes_s1 = [axes_s1]
fig_s1.patch.set_facecolor(DARK_BG)
fig_s1.suptitle("Surface Comparison — Cumulative H & T by Surface Type",
                fontsize=18, fontweight="bold", color=TEXT_CLR)

for i, surf in enumerate(surface_order):
    ax = axes_s1[i]
    ax.set_facecolor(CARD_BG)
    tosses = surface_tosses[surf]
    cum_h, cum_t = cumulative_ht(tosses)
    x = np.arange(1, len(tosses) + 1)

    ax.plot(x, cum_h, color=BLUE, linewidth=2, label="Heads (H)")
    ax.plot(x, cum_t, color=RED, linewidth=2, label="Tails (T)")

    h_tot = cum_h[-1]
    t_tot = cum_t[-1]
    n_tot = len(tosses)

    ax.set_title(f"{surf}\n({n_tot} tosses)", fontsize=13, fontweight="bold",
                 color=SURFACE_COLORS[surf])
    ax.set_xlabel("Toss #", color=TEXT_CLR, fontsize=10)
    ax.set_ylabel("Cumulative Count", color=TEXT_CLR, fontsize=10)
    ax.tick_params(colors=TEXT_CLR, labelsize=9)
    ax.legend(fontsize=10, loc="upper left")
    ax.grid(True, alpha=0.15, color="white")

    ax.annotate(f"H={h_tot} ({h_tot/n_tot*100:.1f}%)",
                xy=(n_tot, h_tot), fontsize=10, color=BLUE, fontweight="bold",
                xytext=(5, 2), textcoords="offset points")
    ax.annotate(f"T={t_tot} ({t_tot/n_tot*100:.1f}%)",
                xy=(n_tot, t_tot), fontsize=10, color=RED, fontweight="bold",
                xytext=(5, -12), textcoords="offset points")

plt.tight_layout(rect=[0, 0, 1, 0.92])
plt.savefig("surface_comparison_cumulative.png", dpi=150, bbox_inches="tight",
            facecolor=fig_s1.get_facecolor())
# plt.show()

print("\n✅  Saved: surface_comparison_cumulative.png")


# ── Chart B: H% bar chart comparison across surfaces ──────────────────────
fig_s2, ax_s2 = plt.subplots(figsize=(10, 7))
fig_s2.patch.set_facecolor(DARK_BG)
ax_s2.set_facecolor(CARD_BG)
fig_s2.suptitle("Surface Comparison — Heads % by Surface Type",
                fontsize=18, fontweight="bold", color=TEXT_CLR)

surfaces = []
h_pcts = []
h_counts = []
t_counts = []
n_counts = []
for surf in surface_order:
    t = surface_tosses[surf]
    h = sum(t)
    surfaces.append(surf)
    h_pcts.append(h / len(t) * 100)
    h_counts.append(h)
    t_counts.append(len(t) - h)
    n_counts.append(len(t))

x_pos = np.arange(len(surfaces))
colors = [SURFACE_COLORS[s] for s in surfaces]
bars = ax_s2.bar(x_pos, h_pcts, color=colors, width=0.5,
                 edgecolor="white", linewidth=1.5)

for bar, pct, h, t, n in zip(bars, h_pcts, h_counts, t_counts, n_counts):
    ax_s2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
               f"{pct:.1f}%\n(H={h}, T={t}, n={n})",
               ha="center", va="bottom", fontsize=11, fontweight="bold", color=TEXT_CLR)

ax_s2.axhline(y=50, color="#64748B", linestyle="--", linewidth=1.5, alpha=0.7, label="50% (fair)")
ax_s2.set_xticks(x_pos)
ax_s2.set_xticklabels(surfaces, fontsize=12, color=TEXT_CLR)
ax_s2.set_ylabel("Heads %", fontsize=13, color=TEXT_CLR)
ax_s2.set_ylim(0, max(h_pcts) + 10)
ax_s2.tick_params(colors=TEXT_CLR)
ax_s2.legend(fontsize=11)
ax_s2.grid(axis="y", alpha=0.2, color="white")

# Summary table
tbl_s = []
for j in range(len(surfaces)):
    tbl_s.append([surfaces[j], str(h_counts[j]), str(t_counts[j]), str(n_counts[j]), f"{h_pcts[j]:.1f}%"])
tbl_surf = ax_s2.table(cellText=tbl_s,
                       colLabels=["Surface", "Heads", "Tails", "Total", "H %"],
                       cellLoc="center", loc="bottom",
                       bbox=[0.10, -0.32, 0.80, 0.20])
tbl_surf.auto_set_font_size(False)
tbl_surf.set_fontsize(11)
for (row, col), cell in tbl_surf.get_celld().items():
    cell.set_edgecolor("white")
    if row == 0:
        cell.set_facecolor("#0F172A")
        cell.set_text_props(color="white", fontweight="bold")
    else:
        cell.set_facecolor("#1E293B")
        cell.set_text_props(color="white")

plt.subplots_adjust(bottom=0.22)
plt.tight_layout(rect=[0, 0.12, 1, 0.93])
plt.savefig("surface_comparison_bar.png", dpi=150, bbox_inches="tight",
            facecolor=fig_s2.get_facecolor())
plt.show()

print("\n✅  Saved: surface_comparison_bar.png")

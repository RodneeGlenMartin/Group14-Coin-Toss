import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict
from matplotlib.gridspec import GridSpec

# =============================================================================
# Load workbook
# =============================================================================
FILENAME = "2BSCS-A _ Tossed Coin Raw Data.xlsx"
wb = openpyxl.load_workbook(FILENAME, data_only=True)

# =============================================================================
# SURFACE ASSIGNMENTS
# =============================================================================
SURFACE_MAP = {
    "GROUP 1": "Wood", "GROUP 2": "Wood", "GROUP 3": "Wood", "GROUP 4": "Wood",
    "GROUP 5": "Wood", "GROUP 6": "Wood", "GROUP 7": "Wood", "GROUP 8": "Wood",
    "GROUP 9": "Tiles", "GROUP 10": "Tiles", "GROUP 11": "Tiles",
    "GROUP 12": "Tiles", "GROUP 13": "Tiles",
    "GROUP 14": "Tiles",
    "GROUP 15": "Tiles",
}

# =============================================================================
# PARSE — extract individual toss results (H=1, T=0)
# =============================================================================
all_records = []  # (group, coin_class, surface, [1/0 per toss])


def derive_from_cumulative(cum_h_list):
    """Given a list of cumulative-H values, return per-toss H (1/0)."""
    results = []
    prev = 0
    for ch in cum_h_list:
        if ch is None:
            ch = prev
        results.append(1 if int(ch) > prev else 0)
        prev = int(ch)
    return results


def safe_int(v, default=0):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def add(group, coin, tosses):
    all_records.append((group, coin, SURFACE_MAP[group], tosses))


# ── GROUP 1  ── 1B & 2 Peso
ws = wb["GROUP 1"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 1", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 1", "Arabian 2", derive_from_cumulative([rows[i][6] for i in range(2, 102)]))

# ── GROUP 2  ── 1B & 5A
ws = wb["GROUP 2"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 2", "1B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 2", "5A", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 3  ── 1B (new) & 10A (old)
ws = wb["GROUP 3"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 3", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 3", "10A", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# ── GROUP 4  ── 5A & 5B
ws = wb["GROUP 4"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 4", "5A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 4", "5B", derive_from_cumulative([rows[i][8] for i in range(2, 102)]))

# ── GROUP 5  ── 1A & 1B
ws = wb["GROUP 5"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 5", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 5", "1B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 6  ── 5B & 20 Peso
ws = wb["GROUP 6"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 6", "5B", derive_from_cumulative([rows[i][3] for i in range(3, 103)]))
add("GROUP 6", "20 Peso", derive_from_cumulative([rows[i][8] for i in range(3, 103)]))

# ── GROUP 7  ── 5A & 10A
ws = wb["GROUP 7"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 7", "5A", derive_from_cumulative([rows[i][4] for i in range(1, 101)]))
add("GROUP 7", "10A", derive_from_cumulative([rows[i][12] for i in range(1, 101)]))

# ── GROUP 8  ── 1A & 10B
ws = wb["GROUP 8"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 8", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 8", "10B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 9  ── 5B (New) & 1B (New) & 20
ws = wb["GROUP 9"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 9", "5B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 9", "1B", derive_from_cumulative([rows[i][7] for i in range(2, 102)]))
add("GROUP 9", "20 Peso", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 10 ── 5B & 10B
ws = wb["GROUP 10"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 10", "5B", derive_from_cumulative([rows[i][3] for i in range(7, 107)]))
add("GROUP 10", "10B", derive_from_cumulative([rows[i][10] for i in range(7, 107)]))

# ── GROUP 11 ── 1A & 10B
ws = wb["GROUP 11"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 11", "1A", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 11", "10B", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 12 ── 5B (New) & 5A (Old)
ws = wb["GROUP 12"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 12", "5B", derive_from_cumulative([rows[i][4] for i in range(3, 103)]))
add("GROUP 12", "5A", derive_from_cumulative([rows[i][12] for i in range(3, 103)]))

# ── GROUP 13 ── 1A & 10A
ws = wb["GROUP 13"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 13", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 13", "10A", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 14 ── 1A & 20 Peso
ws = wb["GROUP 14"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 14", "1A", [safe_int(rows[i][1]) for i in range(2, 102)])
add("GROUP 14", "20 Peso", [safe_int(rows[i][3]) for i in range(2, 102)])

# ── GROUP 15 ── 1B & 5B
ws = wb["GROUP 15"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 15", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 15", "5B", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# =============================================================================
# Verify totals
# =============================================================================
print("=" * 70)
print("  VERIFICATION — Per-group coin totals")
print("=" * 70)
for grp, coin, surf, tosses in all_records:
    h = sum(tosses)
    t = len(tosses) - h
    print(f"  {grp:>10s} | {coin:>12s} | {surf:<22s} | H={h:3d}  T={t:3d}  n={len(tosses)}")
print()

# =============================================================================
# Normalise coin class names & aggregate
# =============================================================================
def normalise(name):
    n = name.strip().upper()
    for rem in ["(NEW)", "(OLD)", "PESO", "₱"]:
        n = n.replace(rem, "")
    return n.strip()


# By coin class
class_tosses = defaultdict(list)
class_groups = defaultdict(list)
class_surface_tosses = defaultdict(lambda: defaultdict(list))

for grp, coin, surf, tosses in all_records:
    key = normalise(coin)
    class_tosses[key].extend(tosses)
    class_groups[key].append(f"G{grp.split()[-1]}({surf[0]})")
    class_surface_tosses[key][surf].extend(tosses)

sorted_classes = sorted(class_tosses.keys())

# Grand combined
all_tosses_combined = []
for tosses in class_tosses.values():
    all_tosses_combined.extend(tosses)

# By surface
surface_tosses = defaultdict(list)
for _, _, surf, tosses in all_records:
    surface_tosses[surf].extend(tosses)

print("  AGGREGATED BY COIN CLASS")
print("-" * 50)
for cls in sorted_classes:
    t = class_tosses[cls]
    groups_str = ", ".join(class_groups[cls])
    print(f"  {cls:>5s}: H={sum(t):4d}  T={len(t)-sum(t):4d}  n={len(t):4d}  [{groups_str}]")
grand_h = sum(all_tosses_combined)
grand_t = len(all_tosses_combined) - grand_h
print(f"  COMBINED: H={grand_h}  T={grand_t}  n={len(all_tosses_combined)}")
print()
print("  AGGREGATED BY SURFACE")
print("-" * 50)
for surf in sorted(surface_tosses.keys()):
    t = surface_tosses[surf]
    if len(t) == 0:
        continue
    h = sum(t)
    print(f"  {surf:>22s}: H={h:4d}  T={len(t)-h:4d}  n={len(t):4d}  ({h/len(t)*100:.1f}% H)")
print("=" * 70)

# =============================================================================
# Helper
# =============================================================================
def cumulative_ht(tosses):
    cum_h, cum_t = [], []
    h, t = 0, 0
    for v in tosses:
        if v == 1:
            h += 1
        else:
            t += 1
        cum_h.append(h)
        cum_t.append(t)
    return cum_h, cum_t


# ── Colour palette
BLUE = "#2563EB"
RED = "#DC2626"
GREEN = "#16A34A"
AMBER = "#F59E0B"
PURPLE = "#9333EA"
DARK_BG = "#1E293B"
CARD_BG = "#334155"
TEXT_CLR = "white"

SURFACE_COLORS = {"Wood": "#3B82F6", "Tiles": "#F97316"}

# =============================================================================
# STEP 3 — All H & T  (Coin Class) — CUMULATIVE LINE CHARTS
# =============================================================================
n_classes = len(sorted_classes)
ncols = min(5, n_classes)
nrows = (n_classes + ncols - 1) // ncols

fig3, axes3 = plt.subplots(nrows, ncols, figsize=(5 * ncols, 5.5 * nrows))
fig3.patch.set_facecolor(DARK_BG)
fig3.suptitle("Step 3 — All H & T (Coin Class) — Cumulative",
              fontsize=18, fontweight="bold", color=TEXT_CLR)

axes_flat = np.array(axes3).flatten()
for idx, cls in enumerate(sorted_classes):
    ax = axes_flat[idx]
    ax.set_facecolor(CARD_BG)
    tosses = class_tosses[cls]
    cum_h, cum_t = cumulative_ht(tosses)
    x = np.arange(1, len(tosses) + 1)

    ax.plot(x, cum_h, color=BLUE, linewidth=1.8, label="Heads (H)")
    ax.plot(x, cum_t, color=RED, linewidth=1.8, label="Tails (T)")

    label = f"₱{cls}" if cls.replace(" ", "").isdigit() else cls
    ax.set_title(label, fontsize=12, fontweight="bold", color=TEXT_CLR)
    ax.set_xlabel("Toss #", color=TEXT_CLR, fontsize=9)
    ax.set_ylabel("Cumulative Count", color=TEXT_CLR, fontsize=9)
    ax.tick_params(colors=TEXT_CLR, labelsize=8)
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(True, alpha=0.15, color="white")

    ax.annotate(f"H={cum_h[-1]}", xy=(len(tosses), cum_h[-1]),
                fontsize=9, color=BLUE, fontweight="bold",
                xytext=(5, 0), textcoords="offset points")
    ax.annotate(f"T={cum_t[-1]}", xy=(len(tosses), cum_t[-1]),
                fontsize=9, color=RED, fontweight="bold",
                xytext=(5, 0), textcoords="offset points")

    # Surface annotation
    groups_str = ", ".join(class_groups[cls])
    ax.text(0.5, -0.18, groups_str, transform=ax.transAxes,
            fontsize=7, color="#94A3B8", ha="center", style="italic")

for j in range(idx + 1, len(axes_flat)):
    axes_flat[j].set_visible(False)

plt.tight_layout(rect=[0, 0.02, 1, 0.94])
plt.savefig("step3_all_ht_coin_class.png", dpi=150, bbox_inches="tight",
            facecolor=fig3.get_facecolor())
# plt.show()

# =============================================================================
# STEP 4 — All H & T  (Combined) — CUMULATIVE LINE CHART
# =============================================================================
fig4, ax4 = plt.subplots(figsize=(12, 8))
fig4.patch.set_facecolor(DARK_BG)
ax4.set_facecolor(CARD_BG)
fig4.suptitle("Step 4 — All H & T (Combined) — Cumulative",
              fontsize=18, fontweight="bold", color=TEXT_CLR)

cum_h_all, cum_t_all = cumulative_ht(all_tosses_combined)
x_all = np.arange(1, len(all_tosses_combined) + 1)

ax4.plot(x_all, cum_h_all, color=BLUE, linewidth=2, label="Heads (H)")
ax4.plot(x_all, cum_t_all, color=RED, linewidth=2, label="Tails (T)")
ax4.set_xlabel("Toss # (all coins concatenated)", fontsize=12, color=TEXT_CLR)
ax4.set_ylabel("Cumulative Count", fontsize=12, color=TEXT_CLR)
ax4.tick_params(colors=TEXT_CLR)
ax4.legend(fontsize=12)
ax4.grid(True, alpha=0.15, color="white")

ax4.annotate(f"H = {cum_h_all[-1]}", xy=(len(all_tosses_combined), cum_h_all[-1]),
             fontsize=12, color=BLUE, fontweight="bold",
             xytext=(5, 5), textcoords="offset points")
ax4.annotate(f"T = {cum_t_all[-1]}", xy=(len(all_tosses_combined), cum_t_all[-1]),
             fontsize=12, color=RED, fontweight="bold",
             xytext=(5, -15), textcoords="offset points")

# Surface legend annotation
surf_summary = "  |  ".join(
    f"{s}: {sum(surface_tosses[s])}/{len(surface_tosses[s])} tosses"
    for s in ["Wood", "Tiles"]
)
fig4.text(0.5, 0.01, f"Surfaces:  {surf_summary}",
          fontsize=9, color="#94A3B8", ha="center", style="italic")

plt.tight_layout(rect=[0, 0.04, 1, 0.93])
plt.savefig("step4_all_ht_combined.png", dpi=150, bbox_inches="tight",
            facecolor=fig4.get_facecolor())
# plt.show()

# =============================================================================
# STEP 5 — Canvas H & T (Coin Class) — Cumulative + Table & Titles
# =============================================================================
n_cls = len(sorted_classes)
ncols = min(n_cls, 5)
grid_rows = (n_cls + ncols - 1) // ncols

gs = GridSpec(grid_rows * 2, ncols, height_ratios=[4, 2] * grid_rows,
              hspace=0.45, wspace=0.35)

fig5 = plt.figure(figsize=(5 * ncols, 5.5 * grid_rows))
fig5.patch.set_facecolor(DARK_BG)
fig5.suptitle("Step 5 — Canvas H & T (Coin Class) — Wood & Tiles",
              fontsize=18, fontweight="bold", color=TEXT_CLR, y=0.98)

for idx, cls in enumerate(sorted_classes):
    col = idx % ncols
    vrow = idx // ncols
    chart_row = vrow * 2
    table_row = vrow * 2 + 1

    # ── Chart subplot
    ax = fig5.add_subplot(gs[chart_row, col])
    ax.set_facecolor(CARD_BG)
    tosses = class_tosses[cls]
    cum_h, cum_t = cumulative_ht(tosses)
    x = np.arange(1, len(tosses) + 1)
    h_total = cum_h[-1]
    t_total = cum_t[-1]
    total = len(tosses)

    ax.plot(x, cum_h, color=BLUE, linewidth=1.8, label="Heads (H)")
    ax.plot(x, cum_t, color=RED, linewidth=1.8, label="Tails (T)")

    label = f"₱{cls}" if cls.replace(" ", "").isdigit() else cls
    ax.set_title(label, fontsize=12, fontweight="bold", color=TEXT_CLR)
    ax.set_xlabel("Toss #", color=TEXT_CLR, fontsize=9)
    ax.set_ylabel("Cumulative Count", color=TEXT_CLR, fontsize=9)
    ax.tick_params(colors=TEXT_CLR, labelsize=8)
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(True, alpha=0.15, color="white")

    ax.annotate(f"H={h_total}", xy=(len(tosses), h_total),
                fontsize=9, color=BLUE, fontweight="bold",
                xytext=(5, 0), textcoords="offset points")
    ax.annotate(f"T={t_total}", xy=(len(tosses), t_total),
                fontsize=9, color=RED, fontweight="bold",
                xytext=(5, 0), textcoords="offset points")

    # ── Table subplot
    ax_tbl = fig5.add_subplot(gs[table_row, col])
    ax_tbl.set_facecolor(DARK_BG)
    ax_tbl.axis("off")

    groups_str = ", ".join(class_groups[cls])

    # Build table data
    table_data = [
        ["Heads (H)", str(h_total), f"{h_total / total * 100:.1f}%"],
        ["Tails (T)", str(t_total), f"{t_total / total * 100:.1f}%"],
        ["Total", str(total), "100.0%"],
        ["", "", ""],
    ]
    surf_data = class_surface_tosses[cls]
    for surf_name in sorted(surf_data.keys()):
        st = surf_data[surf_name]
        sh = sum(st)
        sn = len(st)
        table_data.append([surf_name, f"H={sh} T={sn-sh}", f"{sh/sn*100:.1f}% H"])

    sep_row = 4
    tbl = ax_tbl.table(cellText=table_data,
                       colLabels=["Outcome", "Count", "Probability"],
                       cellLoc="center", loc="upper center",
                       bbox=[0.05, 0.12, 0.90, 0.88])
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    for (row, c), cell in tbl.get_celld().items():
        cell.set_edgecolor("white")
        if row == 0:
            cell.set_facecolor("#0F172A")
            cell.set_text_props(color="white", fontweight="bold")
        elif row == sep_row:
            cell.set_facecolor(DARK_BG)
            cell.set_edgecolor(DARK_BG)
            cell.set_text_props(color=DARK_BG)
        elif row > sep_row:
            cell.set_facecolor("#1E3A5F")
            cell.set_text_props(color="#CBD5E1")
        else:
            cell.set_facecolor("#1E293B")
            cell.set_text_props(color="white")

    # Groups label
    ax_tbl.text(0.5, 0.02, groups_str, transform=ax_tbl.transAxes,
                fontsize=7, color="#94A3B8", ha="center", style="italic")

# Hide unused grid slots
for j in range(n_cls, grid_rows * ncols):
    col = j % ncols
    vrow = j // ncols
    for gr in [vrow * 2, vrow * 2 + 1]:
        ax_empty = fig5.add_subplot(gs[gr, col])
        ax_empty.set_visible(False)

plt.savefig("step5_canvas_ht_coin_class.png", dpi=150, bbox_inches="tight",
            facecolor=fig5.get_facecolor())
# plt.show()

# =============================================================================
# STEP 6 — Canvas H & T (Combined) — Cumulative + Table
# =============================================================================
fig6, ax6 = plt.subplots(figsize=(12, 10))
fig6.patch.set_facecolor(DARK_BG)
ax6.set_facecolor(CARD_BG)
fig6.suptitle("Step 6 — Canvas H & T (Combined)",
              fontsize=18, fontweight="bold", color=TEXT_CLR)

total_all = len(all_tosses_combined)
cum_h_comb, cum_t_comb = cumulative_ht(all_tosses_combined)
x_comb = np.arange(1, total_all + 1)

ax6.plot(x_comb, cum_h_comb, color=BLUE, linewidth=2, label="Heads (H)")
ax6.plot(x_comb, cum_t_comb, color=RED, linewidth=2, label="Tails (T)")
ax6.set_xlabel("Toss # (all coins concatenated)", fontsize=12, color=TEXT_CLR, labelpad=10)
ax6.set_ylabel("Cumulative Count", fontsize=12, color=TEXT_CLR)
ax6.tick_params(colors=TEXT_CLR)
ax6.legend(fontsize=12)
ax6.grid(True, alpha=0.15, color="white")

ax6.annotate(f"H = {grand_h}", xy=(total_all, grand_h),
             fontsize=12, color=BLUE, fontweight="bold",
             xytext=(5, 5), textcoords="offset points")
ax6.annotate(f"T = {grand_t}", xy=(total_all, grand_t),
             fontsize=12, color=RED, fontweight="bold",
             xytext=(5, -15), textcoords="offset points")

# Summary table with surface breakdown -- built dynamically
table_data6 = [
    ["Heads (H)", str(grand_h), f"{grand_h / total_all * 100:.1f}%"],
    ["Tails (T)", str(grand_t), f"{grand_t / total_all * 100:.1f}%"],
    ["Total", str(total_all), "100.0%"],
    ["", "", ""],
]
for surf in sorted(surface_tosses.keys()):
    st = surface_tosses[surf]
    if len(st) == 0:
        continue
    sh = sum(st)
    sn = len(st)
    table_data6.append([surf, f"H={sh} T={sn-sh}", f"{sh/sn*100:.1f}% H"])
tbl6 = ax6.table(cellText=table_data6,
                 colLabels=["Outcome", "Count", "Probability"],
                 cellLoc="center", loc="bottom",
                 bbox=[0.20, -0.55, 0.60, 0.35])
tbl6.auto_set_font_size(False)
tbl6.set_fontsize(10)
for (row, col), cell in tbl6.get_celld().items():
    cell.set_edgecolor("white")
    if row == 0:
        cell.set_facecolor("#0F172A")
        cell.set_text_props(color="white", fontweight="bold")
    elif row == 4:  # separator
        cell.set_facecolor(DARK_BG)
        cell.set_edgecolor(DARK_BG)
        cell.set_text_props(color=DARK_BG)
    elif row >= 5:  # surface rows
        cell.set_facecolor("#1E3A5F")
        cell.set_text_props(color="#CBD5E1")
    else:
        cell.set_facecolor("#1E293B")
        cell.set_text_props(color="white")

plt.subplots_adjust(bottom=0.32)
plt.tight_layout(rect=[0, 0.20, 1, 0.93])
plt.savefig("step6_canvas_ht_combined.png", dpi=150, bbox_inches="tight",
            facecolor=fig6.get_facecolor())
# plt.show()

# =============================================================================
# SURFACE COMPARISON CHARTS
# =============================================================================

# ── Chart A: Cumulative H & T per surface ─────────────────────────────────
surface_order = sorted([s for s in surface_tosses if len(surface_tosses[s]) > 0])
n_surfaces = len(surface_order)
fig_s1, axes_s1 = plt.subplots(1, n_surfaces, figsize=(9 * n_surfaces, 6))
if n_surfaces == 1:
    axes_s1 = [axes_s1]
fig_s1.patch.set_facecolor(DARK_BG)
fig_s1.suptitle("Surface Comparison — Cumulative H & T by Surface Type",
                fontsize=18, fontweight="bold", color=TEXT_CLR)

for i, surf in enumerate(surface_order):
    ax = axes_s1[i]
    ax.set_facecolor(CARD_BG)
    tosses = surface_tosses[surf]
    cum_h, cum_t = cumulative_ht(tosses)
    x = np.arange(1, len(tosses) + 1)

    ax.plot(x, cum_h, color=BLUE, linewidth=2, label="Heads (H)")
    ax.plot(x, cum_t, color=RED, linewidth=2, label="Tails (T)")

    h_tot = cum_h[-1]
    t_tot = cum_t[-1]
    n_tot = len(tosses)

    ax.set_title(f"{surf}\n({n_tot} tosses)", fontsize=13, fontweight="bold",
                 color=SURFACE_COLORS[surf])
    ax.set_xlabel("Toss #", color=TEXT_CLR, fontsize=10)
    ax.set_ylabel("Cumulative Count", color=TEXT_CLR, fontsize=10)
    ax.tick_params(colors=TEXT_CLR, labelsize=9)
    ax.legend(fontsize=10, loc="upper left")
    ax.grid(True, alpha=0.15, color="white")

    ax.annotate(f"H={h_tot} ({h_tot/n_tot*100:.1f}%)",
                xy=(n_tot, h_tot), fontsize=10, color=BLUE, fontweight="bold",
                xytext=(5, 2), textcoords="offset points")
    ax.annotate(f"T={t_tot} ({t_tot/n_tot*100:.1f}%)",
                xy=(n_tot, t_tot), fontsize=10, color=RED, fontweight="bold",
                xytext=(5, -12), textcoords="offset points")

plt.tight_layout(rect=[0, 0, 1, 0.92])
plt.savefig("surface_comparison_cumulative.png", dpi=150, bbox_inches="tight",
            facecolor=fig_s1.get_facecolor())
# plt.show()

# ── Chart B: H% bar chart comparison across surfaces ──────────────────────
fig_s2, ax_s2 = plt.subplots(figsize=(10, 7))
fig_s2.patch.set_facecolor(DARK_BG)
ax_s2.set_facecolor(CARD_BG)
fig_s2.suptitle("Surface Comparison — Heads % by Surface Type",
                fontsize=18, fontweight="bold", color=TEXT_CLR)

surfaces = []
h_pcts = []
h_counts = []
t_counts = []
n_counts = []
for surf in surface_order:
    t = surface_tosses[surf]
    h = sum(t)
    surfaces.append(surf)
    h_pcts.append(h / len(t) * 100)
    h_counts.append(h)
    t_counts.append(len(t) - h)
    n_counts.append(len(t))

x_pos = np.arange(len(surfaces))
colors = [SURFACE_COLORS[s] for s in surfaces]
bars = ax_s2.bar(x_pos, h_pcts, color=colors, width=0.5,
                 edgecolor="white", linewidth=1.5)

for bar, pct, h, t, n in zip(bars, h_pcts, h_counts, t_counts, n_counts):
    ax_s2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
               f"{pct:.1f}%\n(H={h}, T={t}, n={n})",
               ha="center", va="bottom", fontsize=11, fontweight="bold", color=TEXT_CLR)

ax_s2.axhline(y=50, color="#64748B", linestyle="--", linewidth=1.5, alpha=0.7, label="50% (fair)")
ax_s2.set_xticks(x_pos)
ax_s2.set_xticklabels(surfaces, fontsize=12, color=TEXT_CLR)
ax_s2.set_ylabel("Heads %", fontsize=13, color=TEXT_CLR)
ax_s2.set_ylim(0, max(h_pcts) + 10)
ax_s2.tick_params(colors=TEXT_CLR)
ax_s2.legend(fontsize=11)
ax_s2.grid(axis="y", alpha=0.2, color="white")

# Summary table
tbl_s = []
for j in range(len(surfaces)):
    tbl_s.append([surfaces[j], str(h_counts[j]), str(t_counts[j]), str(n_counts[j]), f"{h_pcts[j]:.1f}%"])
tbl_surf = ax_s2.table(cellText=tbl_s,
                       colLabels=["Surface", "Heads", "Tails", "Total", "H %"],
                       cellLoc="center", loc="bottom",
                       bbox=[0.10, -0.32, 0.80, 0.20])
tbl_surf.auto_set_font_size(False)
tbl_surf.set_fontsize(11)
for (row, col), cell in tbl_surf.get_celld().items():
    cell.set_edgecolor("white")
    if row == 0:
        cell.set_facecolor("#0F172A")
        cell.set_text_props(color="white", fontweight="bold")
    else:
        cell.set_facecolor("#1E293B")
        cell.set_text_props(color="white")

plt.subplots_adjust(bottom=0.22)
plt.tight_layout(rect=[0, 0.12, 1, 0.93])
plt.savefig("surface_comparison_bar.png", dpi=150, bbox_inches="tight",
            facecolor=fig_s2.get_facecolor())
plt.show()

# =============================================================================
# EXPORT RAW DATA TO EXCEL  (Steps 3–6 + Surface)
# =============================================================================
out_wb = openpyxl.Workbook()

# ── Step 3 — per coin class cumulative data ───────────────────────────────
for cls in sorted_classes:
    tosses = class_tosses[cls]
    cum_h, cum_t = cumulative_ht(tosses)
    ws = out_wb.create_sheet(title=f"Step3_{cls}")
    ws.append(["Toss #", "Result (1=H, 0=T)", "Cumulative H", "Cumulative T",
               "Groups", ", ".join(class_groups[cls])])
    for i, (res, ch, ct) in enumerate(zip(tosses, cum_h, cum_t), start=1):
        ws.append([i, res, ch, ct])

# ── Step 4 — combined cumulative data ────────────────────────────────────
ws_comb = out_wb.create_sheet(title="Step4_Combined")
ws_comb.append(["Toss #", "Result (1=H, 0=T)", "Cumulative H", "Cumulative T"])
cum_h_all_ex, cum_t_all_ex = cumulative_ht(all_tosses_combined)
for i, (res, ch, ct) in enumerate(zip(all_tosses_combined, cum_h_all_ex, cum_t_all_ex), start=1):
    ws_comb.append([i, res, ch, ct])

# ── Step 5 — per coin class summary tables ────────────────────────────────
ws5 = out_wb.create_sheet(title="Step5_Summary")
ws5.append(["Coin Class", "Heads", "Tails", "Total", "P(H)", "P(T)", "Groups"])
for cls in sorted_classes:
    t = class_tosses[cls]
    h = sum(t)
    tail = len(t) - h
    n = len(t)
    ws5.append([cls, h, tail, n, round(h / n, 4), round(tail / n, 4),
                ", ".join(class_groups[cls])])

# ── Step 6 — combined summary table ──────────────────────────────────────
ws6 = out_wb.create_sheet(title="Step6_Summary")
ws6.append(["Category", "Heads", "Tails", "Total", "P(H)", "P(T)"])
ws6.append(["All Combined", grand_h, grand_t, total_all,
            round(grand_h / total_all, 4), round(grand_t / total_all, 4)])
ws6.append([])
ws6.append(["BY SURFACE", "Heads", "Tails", "Total", "P(H)", "P(T)"])
for surf in surface_order:
    t = surface_tosses[surf]
    h = sum(t)
    tail = len(t) - h
    n = len(t)
    ws6.append([surf, h, tail, n, round(h / n, 4), round(tail / n, 4)])

# ── Surface — per surface cumulative data ─────────────────────────────────
for surf in surface_order:
    safe_name = surf.replace(" ", "")[:20]
    tosses = surface_tosses[surf]
    cum_h, cum_t = cumulative_ht(tosses)
    ws_s = out_wb.create_sheet(title=f"Surf_{safe_name}")
    ws_s.append(["Toss #", "Result (1=H, 0=T)", "Cumulative H", "Cumulative T"])
    for i, (res, ch, ct) in enumerate(zip(tosses, cum_h, cum_t), start=1):
        ws_s.append([i, res, ch, ct])

# Remove default empty sheet
if "Sheet" in out_wb.sheetnames:
    del out_wb["Sheet"]

OUT_EXCEL = "steps3_6_raw_data.xlsx"
out_wb.save(OUT_EXCEL)

print(f"\n✅  All charts saved:")
print(f"  • step3_all_ht_coin_class.png")
print(f"  • step4_all_ht_combined.png")
print(f"  • step5_canvas_ht_coin_class.png")
print(f"  • step6_canvas_ht_combined.png")
print(f"  • surface_comparison_cumulative.png")
print(f"  • surface_comparison_bar.png")
print(f"\n📊  Raw data exported to:  {OUT_EXCEL}")
print(f"    Sheets: {out_wb.sheetnames}")

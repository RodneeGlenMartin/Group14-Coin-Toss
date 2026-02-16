import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict

# =============================================================================
# Load workbook
# =============================================================================
FILENAME = "2BSCS-A _ Tossed Coin Raw Data.xlsx"
wb = openpyxl.load_workbook(FILENAME, data_only=True)

# =============================================================================
# SURFACE ASSIGNMENTS
# =============================================================================
SURFACE_MAP = {
    "GROUP 1": "Wood", "GROUP 2": "Wood", "GROUP 3": "Wood", "GROUP 4": "Wood",
    "GROUP 5": "Wood", "GROUP 6": "Wood", "GROUP 7": "Wood", "GROUP 8": "Wood",
    "GROUP 9": "Tiles", "GROUP 10": "Tiles", "GROUP 11": "Tiles",
    "GROUP 12": "Tiles", "GROUP 13": "Tiles",
    "GROUP 14": "Tiles",
    "GROUP 15": "Tiles",
}

# =============================================================================
# PARSE — extract individual toss results (H=1, T=0)
# =============================================================================
all_records = []


def derive_from_cumulative(cum_h_list):
    """Given a list of cumulative-H values, return per-toss H (1/0)."""
    results = []
    prev = 0
    for ch in cum_h_list:
        if ch is None:
            ch = prev
        results.append(1 if int(ch) > prev else 0)
        prev = int(ch)
    return results


def safe_int(v, default=0):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def add(group, coin, tosses):
    all_records.append((group, coin, SURFACE_MAP[group], tosses))


# ── GROUP 1  ── 1B & 2 Peso
ws = wb["GROUP 1"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 1", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 1", "Arabian 2", derive_from_cumulative([rows[i][6] for i in range(2, 102)]))

# ── GROUP 2  ── 1B & 5A
ws = wb["GROUP 2"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 2", "1B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 2", "5A", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 3  ── 1B (new) & 10A (old)
ws = wb["GROUP 3"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 3", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 3", "10A", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# ── GROUP 4  ── 5A & 5B
ws = wb["GROUP 4"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 4", "5A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 4", "5B", derive_from_cumulative([rows[i][8] for i in range(2, 102)]))

# ── GROUP 5  ── 1A & 1B
ws = wb["GROUP 5"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 5", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 5", "1B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 6  ── 5B & 20 Peso
ws = wb["GROUP 6"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 6", "5B", derive_from_cumulative([rows[i][3] for i in range(3, 103)]))
add("GROUP 6", "20 Peso", derive_from_cumulative([rows[i][8] for i in range(3, 103)]))

# ── GROUP 7  ── 5A & 10A
ws = wb["GROUP 7"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 7", "5A", derive_from_cumulative([rows[i][4] for i in range(1, 101)]))
add("GROUP 7", "10A", derive_from_cumulative([rows[i][12] for i in range(1, 101)]))

# ── GROUP 8  ── 1A & 10B
ws = wb["GROUP 8"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 8", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 8", "10B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

# ── GROUP 9  ── 5B (New) & 1B (New) & 20
ws = wb["GROUP 9"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 9", "5B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 9", "1B", derive_from_cumulative([rows[i][7] for i in range(2, 102)]))
add("GROUP 9", "20 Peso", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

# ── GROUP 10 ── 5B & 10B
ws = wb["GROUP 10"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 10", "5B", derive_from_cumulative([rows[i][3] for i in range(7, 107)]))
add("GROUP 10", "10B", derive_from_cumulative([rows[i][10] for i in range(7, 107)]))

# ── GROUP 11 ── 1A & 10B
ws = wb["GROUP 11"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 11", "1A", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 11", "10B", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 12 ── 5B (New) & 5A (Old)
ws = wb["GROUP 12"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 12", "5B", derive_from_cumulative([rows[i][4] for i in range(3, 103)]))
add("GROUP 12", "5A", derive_from_cumulative([rows[i][12] for i in range(3, 103)]))

# ── GROUP 13 ── 1A & 10A
ws = wb["GROUP 13"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 13", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 13", "10A", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

# ── GROUP 14 ── 1A & 20 Peso
ws = wb["GROUP 14"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 14", "1A", [safe_int(rows[i][1]) for i in range(2, 102)])
add("GROUP 14", "20 Peso", [safe_int(rows[i][3]) for i in range(2, 102)])

# ── GROUP 15 ── 1B & 5B
ws = wb["GROUP 15"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 15", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 15", "5B", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# =============================================================================
# Normalise coin class names & aggregate
# =============================================================================
def normalise(name):
    n = name.strip().upper()
    for rem in ["(NEW)", "(OLD)", "PESO", "₱"]:
        n = n.replace(rem, "")
    return n.strip()


class_tosses = defaultdict(list)
class_groups = defaultdict(list)

# Also track per coin-class per surface
class_surface_tosses = defaultdict(lambda: defaultdict(list))

for grp, coin, surf, tosses in all_records:
    key = normalise(coin)
    class_tosses[key].extend(tosses)
    class_groups[key].append(f"G{grp.split()[-1]}({surf[0]})")
    class_surface_tosses[key][surf].extend(tosses)

sorted_classes = sorted(class_tosses.keys())

# =============================================================================
# Helper
# =============================================================================
def cumulative_ht(tosses):
    cum_h, cum_t = [], []
    h, t = 0, 0
    for v in tosses:
        if v == 1:
            h += 1
        else:
            t += 1
        cum_h.append(h)
        cum_t.append(t)
    return cum_h, cum_t


# ── Colour palette
BLUE = "#2563EB"
RED = "#DC2626"
DARK_BG = "#1E293B"
CARD_BG = "#334155"
TEXT_CLR = "white"

# =============================================================================
# STEP 5 — Canvas H & T (Coin Class) — Cumulative + Table & Surface
# =============================================================================
from matplotlib.gridspec import GridSpec

n_cls = len(sorted_classes)
ncols = min(n_cls, 5)
grid_rows = (n_cls + ncols - 1) // ncols  # how many visual rows of panels

# Each visual row gets 2 grid rows: chart (height 4) + table (height 2)
gs = GridSpec(grid_rows * 2, ncols, height_ratios=[4, 2] * grid_rows,
              hspace=0.45, wspace=0.35)

fig5 = plt.figure(figsize=(5 * ncols, 5.5 * grid_rows))
fig5.patch.set_facecolor(DARK_BG)
fig5.suptitle("Step 5 — Canvas H & T (Coin Class) — Wood & Tiles",
              fontsize=18, fontweight="bold", color=TEXT_CLR, y=0.98)

for idx, cls in enumerate(sorted_classes):
    col = idx % ncols
    vrow = idx // ncols
    chart_row = vrow * 2
    table_row = vrow * 2 + 1

    # ── Chart subplot
    ax = fig5.add_subplot(gs[chart_row, col])
    ax.set_facecolor(CARD_BG)
    tosses = class_tosses[cls]
    cum_h, cum_t = cumulative_ht(tosses)
    x = np.arange(1, len(tosses) + 1)
    h_total = cum_h[-1]
    t_total = cum_t[-1]
    total = len(tosses)

    ax.plot(x, cum_h, color=BLUE, linewidth=1.8, label="Heads (H)")
    ax.plot(x, cum_t, color=RED, linewidth=1.8, label="Tails (T)")

    label = f"₱{cls}" if cls.replace(" ", "").isdigit() else cls
    ax.set_title(label, fontsize=12, fontweight="bold", color=TEXT_CLR)
    ax.set_xlabel("Toss #", color=TEXT_CLR, fontsize=9)
    ax.set_ylabel("Cumulative Count", color=TEXT_CLR, fontsize=9)
    ax.tick_params(colors=TEXT_CLR, labelsize=8)
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(True, alpha=0.15, color="white")

    ax.annotate(f"H={h_total}", xy=(len(tosses), h_total),
                fontsize=9, color=BLUE, fontweight="bold",
                xytext=(5, 0), textcoords="offset points")
    ax.annotate(f"T={t_total}", xy=(len(tosses), t_total),
                fontsize=9, color=RED, fontweight="bold",
                xytext=(5, 0), textcoords="offset points")

    # ── Table subplot
    ax_tbl = fig5.add_subplot(gs[table_row, col])
    ax_tbl.set_facecolor(DARK_BG)
    ax_tbl.axis("off")

    groups_str = ", ".join(class_groups[cls])

    # Build table data
    table_data = [
        ["Heads (H)", str(h_total), f"{h_total / total * 100:.1f}%"],
        ["Tails (T)", str(t_total), f"{t_total / total * 100:.1f}%"],
        ["Total", str(total), "100.0%"],
        ["", "", ""],  # separator
    ]
    surf_data = class_surface_tosses[cls]
    for surf_name in sorted(surf_data.keys()):
        st = surf_data[surf_name]
        sh = sum(st)
        sn = len(st)
        table_data.append([surf_name, f"H={sh} T={sn-sh}", f"{sh/sn*100:.1f}% H"])

    n_data_rows = len(table_data)
    sep_row = 4  # 1-indexed; row 0 = header

    tbl = ax_tbl.table(cellText=table_data,
                       colLabels=["Outcome", "Count", "Probability"],
                       cellLoc="center", loc="upper center",
                       bbox=[0.05, 0.12, 0.90, 0.88])
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    for (row, c), cell in tbl.get_celld().items():
        cell.set_edgecolor("white")
        if row == 0:
            cell.set_facecolor("#0F172A")
            cell.set_text_props(color="white", fontweight="bold")
        elif row == sep_row:
            cell.set_facecolor(DARK_BG)
            cell.set_edgecolor(DARK_BG)
            cell.set_text_props(color=DARK_BG)
        elif row > sep_row:
            cell.set_facecolor("#1E3A5F")
            cell.set_text_props(color="#CBD5E1")
        else:
            cell.set_facecolor("#1E293B")
            cell.set_text_props(color="white")

    # Groups label at bottom of table area
    ax_tbl.text(0.5, 0.02, groups_str, transform=ax_tbl.transAxes,
                fontsize=7, color="#94A3B8", ha="center", style="italic")

# Hide unused grid slots
for j in range(n_cls, grid_rows * ncols):
    col = j % ncols
    vrow = j // ncols
    for gr in [vrow * 2, vrow * 2 + 1]:
        ax_empty = fig5.add_subplot(gs[gr, col])
        ax_empty.set_visible(False)

plt.savefig("group14_step5.png", dpi=150, bbox_inches="tight",
            facecolor=fig5.get_facecolor())
# plt.show()

print("\n✅  Step 5 chart saved: step5_canvas_ht_coin_class.png")

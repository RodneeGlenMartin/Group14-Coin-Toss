import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict

FILENAME = "2BSCS-A _ Tossed Coin Raw Data.xlsx"
wb = openpyxl.load_workbook(FILENAME, data_only=True)

SURFACE_MAP = {
    "GROUP 1": "Wood", "GROUP 2": "Wood", "GROUP 3": "Wood", "GROUP 4": "Wood",
    "GROUP 5": "Wood", "GROUP 6": "Wood", "GROUP 7": "Wood", "GROUP 8": "Wood",
    "GROUP 9": "Tiles", "GROUP 10": "Tiles", "GROUP 11": "Tiles",
    "GROUP 12": "Tiles", "GROUP 13": "Tiles",
    "GROUP 14": "Tiles",
    "GROUP 15": "Tiles",
}

all_records = []
def derive_from_cumulative(cum_h_list):
    results = []
    prev = 0
    for ch in cum_h_list:
        if ch is None: ch = prev
        results.append(1 if int(ch) > prev else 0)
        prev = int(ch)
    return results

def safe_int(v, default=0):
    if v is None: return default
    try: return int(v)
    except: return default

def add(group, coin, tosses):
    all_records.append((group, coin, SURFACE_MAP[group], tosses))

ws = wb["GROUP 1"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 1", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 1", "Arabian 2", derive_from_cumulative([rows[i][6] for i in range(2, 102)]))
# ... (Adding only Group 1 for brevity in prompt context? No, must add ALL for Global Step 4)
# I will assume I need ALL groups logic here. Detailed list:

ws = wb["GROUP 2"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 2", "1B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 2", "5A", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

ws = wb["GROUP 3"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 3", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 3", "10A", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

ws = wb["GROUP 4"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 4", "5A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 4", "5B", derive_from_cumulative([rows[i][8] for i in range(2, 102)]))

ws = wb["GROUP 5"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 5", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 5", "1B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

ws = wb["GROUP 6"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 6", "5B", derive_from_cumulative([rows[i][3] for i in range(3, 103)]))
add("GROUP 6", "20 Peso", derive_from_cumulative([rows[i][8] for i in range(3, 103)]))

ws = wb["GROUP 7"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 7", "5A", derive_from_cumulative([rows[i][4] for i in range(1, 101)]))
add("GROUP 7", "10A", derive_from_cumulative([rows[i][12] for i in range(1, 101)]))

ws = wb["GROUP 8"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 8", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 8", "10B", derive_from_cumulative([rows[i][9] for i in range(2, 102)]))

ws = wb["GROUP 9"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 9", "5B", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 9", "1B", derive_from_cumulative([rows[i][7] for i in range(2, 102)]))
add("GROUP 9", "20 Peso", derive_from_cumulative([rows[i][11] for i in range(2, 102)]))

ws = wb["GROUP 10"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 10", "5B", derive_from_cumulative([rows[i][3] for i in range(7, 107)]))
add("GROUP 10", "10B", derive_from_cumulative([rows[i][10] for i in range(7, 107)]))

ws = wb["GROUP 11"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 11", "1A", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 11", "10B", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

ws = wb["GROUP 12"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 12", "5B", derive_from_cumulative([rows[i][4] for i in range(3, 103)]))
add("GROUP 12", "5A", derive_from_cumulative([rows[i][12] for i in range(3, 103)]))

ws = wb["GROUP 13"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 13", "1A", derive_from_cumulative([rows[i][3] for i in range(2, 102)]))
add("GROUP 13", "10A", derive_from_cumulative([rows[i][10] for i in range(2, 102)]))

ws = wb["GROUP 14"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 14", "1A", [safe_int(rows[i][1]) for i in range(2, 102)])
add("GROUP 14", "20 Peso", [safe_int(rows[i][3]) for i in range(2, 102)])

ws = wb["GROUP 15"]; rows = list(ws.iter_rows(values_only=True))
add("GROUP 15", "1B", derive_from_cumulative([rows[i][4] for i in range(2, 102)]))
add("GROUP 15", "5B", derive_from_cumulative([rows[i][12] for i in range(2, 102)]))

# Combined tosses of ALL groups
all_tosses_combined = []
surface_tosses = defaultdict(list)
for grp, coin, surf, tosses in all_records:
    all_tosses_combined.extend(tosses)
    surface_tosses[surf].extend(tosses)

def cumulative_ht(tosses):
    cum_h, cum_t = [], []
    h, t = 0, 0
    for v in tosses:
        if v == 1: h += 1
        else: t += 1
        cum_h.append(h)
        cum_t.append(t)
    return cum_h, cum_t

BLUE = "#2563EB"
RED = "#DC2626"
DARK_BG = "#1E293B"
CARD_BG = "#334155"
TEXT_CLR = "white"

fig4, ax4 = plt.subplots(figsize=(12, 8))
fig4.patch.set_facecolor(DARK_BG)
ax4.set_facecolor(CARD_BG)
fig4.suptitle("Step 4 — All H & T (Combined) — Cumulative", fontsize=18, fontweight="bold", color=TEXT_CLR)

cum_h_all, cum_t_all = cumulative_ht(all_tosses_combined)
x_all = np.arange(1, len(all_tosses_combined) + 1)

ax4.plot(x_all, cum_h_all, color=BLUE, linewidth=2, label="Heads (H)")
ax4.plot(x_all, cum_t_all, color=RED, linewidth=2, label="Tails (T)")
ax4.set_xlabel("Toss # (all coins concatenated)", fontsize=12, color=TEXT_CLR)
ax4.set_ylabel("Cumulative Count", fontsize=12, color=TEXT_CLR)
ax4.tick_params(colors=TEXT_CLR)
ax4.legend(fontsize=12)
ax4.grid(True, alpha=0.15, color="white")

ax4.annotate(f"H = {cum_h_all[-1]}", xy=(len(all_tosses_combined), cum_h_all[-1]), fontsize=12, color=BLUE, fontweight="bold", xytext=(5, 5), textcoords="offset points")
ax4.annotate(f"T = {cum_t_all[-1]}", xy=(len(all_tosses_combined), cum_t_all[-1]), fontsize=12, color=RED, fontweight="bold", xytext=(5, -15), textcoords="offset points")

surf_summary = "  |  ".join(
    f"{s}: {sum(surface_tosses[s])}/{len(surface_tosses[s])} tosses"
    for s in ["Wood", "Tiles"]
)
fig4.text(0.5, 0.01, f"Surfaces:  {surf_summary}", fontsize=9, color="#94A3B8", ha="center", style="italic")

plt.tight_layout(rect=[0, 0.04, 1, 0.93])
plt.savefig("group14_step4.png", facecolor=fig4.get_facecolor())
print("Saved group14_step4.png")
